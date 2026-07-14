from __future__ import annotations

from dataclasses import dataclass
from enum import StrEnum
from io import BytesIO
from pathlib import Path

import pandas as pd
from openpyxl import load_workbook

from app.core.errors import IngestError


class FileFormat(StrEnum):
    CSV = "csv"
    XLSX = "xlsx"
    XLS = "xls"


@dataclass(frozen=True, slots=True)
class TableArtifact:
    """One loadable table: a CSV file or a single sheet of a workbook (T58)."""

    label: str
    source_filename: str
    sheet_name: str | None
    frame: pd.DataFrame

    @property
    def headers(self) -> list[str]:
        return [str(column).strip() for column in self.frame.columns]


def detect_file_format(file_name: str) -> FileFormat:
    suffix = Path(file_name).suffix.lower().lstrip(".")
    if suffix in {FileFormat.CSV, FileFormat.XLSX, FileFormat.XLS}:
        return FileFormat(suffix)
    msg = f"Unsupported file extension for {file_name}"
    raise IngestError(msg)


def load_table(
    file_name: str,
    content: bytes,
    *,
    sheet_name: str | None = None,
) -> pd.DataFrame:
    format_ = detect_file_format(file_name)
    if format_ == FileFormat.CSV:
        if sheet_name:
            msg = f"CSV upload cannot select sheet {sheet_name!r}"
            raise IngestError(msg)
        return pd.read_csv(BytesIO(content))
    if format_ in {FileFormat.XLSX, FileFormat.XLS}:
        return load_excel(content, sheet_name=sheet_name)
    msg = f"Unsupported file extension for {file_name}"
    raise IngestError(msg)


def list_excel_sheet_names(content: bytes) -> list[str]:
    workbook = load_workbook(BytesIO(content), data_only=True, read_only=True)
    try:
        return list(workbook.sheetnames)
    finally:
        workbook.close()


def load_excel(content: bytes, *, sheet_name: str | None = None) -> pd.DataFrame:
    workbook = load_workbook(BytesIO(content), data_only=True, read_only=True)
    try:
        if sheet_name is None:
            worksheet = workbook.active
            if worksheet is None:
                msg = "Workbook does not contain an active worksheet"
                raise IngestError(msg)
        else:
            if sheet_name not in workbook.sheetnames:
                msg = f"Worksheet not found: {sheet_name}"
                raise IngestError(msg)
            worksheet = workbook[sheet_name]
        rows = list(worksheet.iter_rows(values_only=True))
    finally:
        workbook.close()
    if not rows:
        return pd.DataFrame()
    headers = ["" if header is None else str(header).strip() for header in rows[0]]
    return pd.DataFrame(rows[1:], columns=headers)


def expand_upload_tables(file_name: str, content: bytes) -> list[TableArtifact]:
    """Expand one upload into one or more table artifacts (multi-sheet xlsx → N)."""
    format_ = detect_file_format(file_name)
    if format_ == FileFormat.CSV:
        frame = pd.read_csv(BytesIO(content))
        return [
            TableArtifact(
                label=file_name,
                source_filename=file_name,
                sheet_name=None,
                frame=frame,
            )
        ]
    if format_ not in {FileFormat.XLSX, FileFormat.XLS}:
        msg = f"Unsupported file extension for {file_name}"
        raise IngestError(msg)

    sheet_names = list_excel_sheet_names(content)
    if not sheet_names:
        msg = f"Workbook has no worksheets: {file_name}"
        raise IngestError(msg)
    if len(sheet_names) == 1:
        only = sheet_names[0]
        frame = load_excel(content, sheet_name=only)
        return [
            TableArtifact(
                label=file_name,
                source_filename=file_name,
                sheet_name=only,
                frame=frame,
            )
        ]
    artifacts: list[TableArtifact] = []
    for name in sheet_names:
        frame = load_excel(content, sheet_name=name)
        artifacts.append(
            TableArtifact(
                label=f"{file_name} · {name}",
                source_filename=file_name,
                sheet_name=name,
                frame=frame,
            )
        )
    return artifacts
